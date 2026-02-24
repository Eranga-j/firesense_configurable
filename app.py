from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
import bcrypt

# For Excel reports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export disabled. Install with: pip install openpyxl")

# For PDF reports  
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️  reportlab not installed. PDF export disabled. Install with: pip install reportlab")

app = Flask(__name__)
app.config['SECRET_KEY'] = 'firesense-configurable-2025'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:postgres123@localhost/firesense_hotel'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==================== DATABASE MODELS ====================

class HotelSettings(db.Model):
    __tablename__ = 'hotel_settings'
    id = db.Column(db.Integer, primary_key=True)
    hotel_name = db.Column(db.String(200), nullable=False)
    city = db.Column(db.String(100))
    province = db.Column(db.String(100))
    address = db.Column(db.String(500))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    website = db.Column(db.String(200))
    star_rating = db.Column(db.Integer)
    total_rooms = db.Column(db.Integer)
    hotel_type = db.Column(db.String(100))
    logo_url = db.Column(db.String(500))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = db.Column(db.Integer, db.ForeignKey('users.id'))

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(20))
    date_joined = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='ACTIVE')

class Equipment(db.Model):
    __tablename__ = 'equipment'
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.String(50), unique=True)
    equipment_type = db.Column(db.String(100), nullable=False)
    serial_number = db.Column(db.String(100))
    manufacturer = db.Column(db.String(100))
    model = db.Column(db.String(100))
    install_year = db.Column(db.Integer, nullable=False)
    install_date = db.Column(db.Date)
    last_service_date = db.Column(db.Date, nullable=False)
    next_service_date = db.Column(db.Date)
    service_interval_months = db.Column(db.Integer, default=6)
    location = db.Column(db.String(200))
    floor = db.Column(db.String(50))
    zone = db.Column(db.String(100))
    near_emergency_exit = db.Column(db.Boolean, default=False)
    condition = db.Column(db.String(50), default='Good')
    status = db.Column(db.String(20), default='ACTIVE')
    usage_level = db.Column(db.String(20))
    daily_exposure_hours = db.Column(db.Integer)
    humidity_level = db.Column(db.String(20))
    temperature_avg = db.Column(db.Integer)
    coastal_exposure = db.Column(db.Boolean, default=False)
    certification_status = db.Column(db.String(50), default='Valid')
    last_audit_score = db.Column(db.Integer)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))

class MaintenanceRecord(db.Model):
    __tablename__ = 'maintenance_record'
    id = db.Column(db.Integer, primary_key=True)
    record_id = db.Column(db.String(50), unique=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    maintenance_date = db.Column(db.Date, nullable=False)
    maintenance_type = db.Column(db.String(50))
    technician_name = db.Column(db.String(100))
    technician_company = db.Column(db.String(200))
    issues_found = db.Column(db.Text)
    actions_taken = db.Column(db.Text)
    parts_replaced = db.Column(db.Text)
    cost_lkr = db.Column(db.Float)
    duration_hours = db.Column(db.Float)
    next_service_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='Completed')
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    equipment = db.relationship('Equipment', backref='maintenance_records')
    user = db.relationship('User', backref='maintenance_records')

class Prediction(db.Model):
    __tablename__ = 'predictions'
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    risk_level = db.Column(db.String(20), nullable=False)
    risk_score = db.Column(db.Integer)
    prediction_date = db.Column(db.DateTime, default=datetime.utcnow)
    summary = db.Column(db.Text)
    status = db.Column(db.String(20), default='ACTIVE')
    equipment = db.relationship('Equipment', backref='predictions')
    user = db.relationship('User', backref='predictions')

class Alert(db.Model):
    __tablename__ = 'alert'
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'))
    alert_type = db.Column(db.String(50))
    severity = db.Column(db.String(20))
    message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    is_resolved = db.Column(db.Boolean, default=False)
    resolved_at = db.Column(db.DateTime)
    resolved_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    read_at = db.Column(db.DateTime)
    read_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    equipment = db.relationship('Equipment', backref='alerts')

# ==================== CHAT SYSTEM MODELS ====================

class ChatRoom(db.Model):
    __tablename__ = 'chat_rooms'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    room_type = db.Column(db.String(50))  # 'general', 'equipment', 'direct'
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationships
    messages = db.relationship('ChatMessage', backref='room', lazy='dynamic', cascade='all, delete-orphan')
    equipment = db.relationship('Equipment', backref='chat_rooms')

class ChatMessage(db.Model):
    __tablename__ = 'chat_messages'
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('chat_rooms.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    attachment_url = db.Column(db.String(500))
    attachment_name = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_edited = db.Column(db.Boolean, default=False)
    edited_at = db.Column(db.DateTime)
    
    # Relationships
    user = db.relationship('User', backref='messages')
    read_by = db.relationship('ChatReadReceipt', backref='message', cascade='all, delete-orphan')

class ChatReadReceipt(db.Model):
    __tablename__ = 'chat_read_receipts'
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.Integer, db.ForeignKey('chat_messages.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    read_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='read_receipts')

# ==================== HELPER FUNCTIONS ====================

def get_hotel_info():
    settings = HotelSettings.query.first()
    if not settings:
        settings = HotelSettings(hotel_name='Your Hotel Name', city='Colombo', province='Western', star_rating=5, total_rooms=100, hotel_type='City Hotel')
        db.session.add(settings)
        db.session.commit()
    return {
        'id': settings.id, 'hotel_name': settings.hotel_name, 'city': settings.city,
        'province': settings.province, 'address': settings.address, 'phone': settings.phone,
        'email': settings.email, 'website': settings.website, 'star_rating': settings.star_rating,
        'total_rooms': settings.total_rooms, 'hotel_type': settings.hotel_type,
        'logo_url': settings.logo_url,
        'updated_at': settings.updated_at.strftime('%Y-%m-%d') if settings.updated_at else None
    }

def predict_risk_simple(equipment):
    risk_score = 0
    age = datetime.now().year - equipment.install_year
    if age > 10: risk_score += 30
    elif age > 5: risk_score += 15
    if equipment.last_service_date and equipment.next_service_date:
        days_overdue = (datetime.now().date() - equipment.next_service_date).days
        if days_overdue > 60: risk_score += 40
        elif days_overdue > 30: risk_score += 25
        elif days_overdue > 0: risk_score += 10
    if equipment.usage_level == 'High': risk_score += 15
    elif equipment.usage_level == 'Medium': risk_score += 8
    if equipment.humidity_level == 'High': risk_score += 5
    if equipment.coastal_exposure: risk_score += 5
    if equipment.condition == 'Poor': risk_score += 10
    elif equipment.condition == 'Fair': risk_score += 5
    if risk_score >= 60: return 'High', risk_score
    elif risk_score >= 30: return 'Medium', risk_score
    else: return 'Low', risk_score

def create_default_chat_rooms():
    """Create default chat rooms if they don't exist"""
    general_chat = ChatRoom.query.filter_by(room_type='general', name='General Discussion').first()
    
    if not general_chat:
        admin = User.query.filter_by(role='Admin').first()
        
        general = ChatRoom(
            name='General Discussion',
            room_type='general',
            created_by=admin.id if admin else 1
        )
        
        maintenance = ChatRoom(
            name='Maintenance Team',
            room_type='general',
            created_by=admin.id if admin else 1
        )
        
        urgent = ChatRoom(
            name='Urgent Issues',
            room_type='general',
            created_by=admin.id if admin else 1
        )
        
        db.session.add(general)
        db.session.add(maintenance)
        db.session.add(urgent)
        db.session.commit()
        
        print("✅ Default chat rooms created!")

# ==================== AUTHENTICATION DECORATORS ====================
def create_default_users():
    """Create default users if they don't exist"""
    admin = User.query.filter_by(username='admin').first()
    
    if not admin:
        admin_password = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        manager_password = bcrypt.hashpw('manager123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        tech_password = bcrypt.hashpw('tech123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        admin_user = User(username='admin', password=admin_password, role='Admin', email='admin@firesense.com', phone='0771234567')
        manager_user = User(username='manager', password=manager_password, role='Manager', email='manager@firesense.com', phone='0771234568')
        tech_user = User(username='tech1', password=tech_password, role='Technician', email='tech1@firesense.com', phone='0771234569')
        
        db.session.add(admin_user)
        db.session.add(manager_user)
        db.session.add(tech_user)
        db.session.commit()
        
        print("✅ Default users created!")
        print("   👑 Admin: admin / admin123")
        print("   👔 Manager: manager / manager123")
        print("   🔧 Tech: tech1 / tech123")
        
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'Admin':
            return jsonify({'success': False, 'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

def admin_or_manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ['Admin', 'Manager']:
            return jsonify({'success': False, 'error': 'Admin or Manager access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# ==================== CONTEXT PROCESSOR ====================

@app.context_processor
def inject_user_info():
    hotel = get_hotel_info()
    current_user_role = session.get('role')
    return {
        'hotel': hotel,
        'current_user_role': current_user_role
    }

# ==================== ROUTES - PAGES ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        print(f"Login attempt: username={username}")
        print(f"User found: {user is not None}")
        if user:
            print(f"Stored hash: {user.password}")
            print(f"Attempting with password: {password}")
        
        if user:
            try:
                if bcrypt.checkpw(password.encode('utf-8'), user.password.encode('utf-8')):
                    session['user_id'] = user.id
                    session['username'] = user.username
                    session['role'] = user.role
                    return jsonify({'success': True})
            except Exception as e:
                print(f"Bcrypt error: {e}")
        
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/equipment-list')
@login_required
def equipment_list():
    return render_template('equipment_list.html')

@app.route('/equipment/<int:equipment_id>')
@login_required
def equipment_details(equipment_id):
    return render_template('equipment_details.html', equipment_id=equipment_id)

@app.route('/add-maintenance/<int:equipment_id>', methods=['GET', 'POST'])
@login_required
def add_maintenance(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    
    if request.method == 'POST':
        maintenance = MaintenanceRecord(
            equipment_id=equipment.id,
            user_id=session.get('user_id'),
            maintenance_date=request.form.get('maintenance_date'),
            maintenance_type=request.form.get('maintenance_type'),
            technician_name=request.form.get('technician_name'),
            technician_company=request.form.get('technician_company'),
            issues_found=request.form.get('issues_found'),
            actions_taken=request.form.get('actions_taken'),
            parts_replaced=request.form.get('parts_replaced'),
            cost_lkr=request.form.get('cost_lkr') or 0,
            duration_hours=request.form.get('duration_hours') or 0,
            next_service_date=request.form.get('next_service_date') or None,  # ← FIXED: Convert empty string to None
            status=request.form.get('status', 'COMPLETED'),
            notes=request.form.get('notes')
        )
        
        equipment.last_service_date = request.form.get('maintenance_date')
        next_service_input = request.form.get('next_service_date')
        if next_service_input and next_service_input.strip():
            equipment.next_service_date = next_service_input
        else:
            equipment.next_service_date = None  # ← FIXED: Set to None if empty
        
        db.session.add(maintenance)
        db.session.commit()
        
        return redirect(url_for('equipment_details', equipment_id=equipment.id))
    
    from datetime import date
    today = date.today().isoformat()
    
    hotel_settings = HotelSettings.query.first()
    hotel_name = hotel_settings.hotel_name if hotel_settings else "FireSense"
    
    return render_template('add_maintenance.html',
                         equipment=equipment,
                         today=today,
                         hotel_name=hotel_name,
                         current_user=session.get('username'),
                         current_user_role=session.get('role'))

@app.route('/prediction-history')
@login_required
def prediction_history():
    return render_template('prediction_history.html')

@app.route('/user-management')
@login_required
@admin_or_manager_required
def user_management():
    return render_template('user_management.html')

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')

@app.route('/alerts')
@login_required
def alerts_page():
    return render_template('alerts.html')

@app.route('/settings')
@login_required
@admin_required
def settings_page():
    return render_template('settings.html')

@app.route('/equipment/add')
@login_required
@admin_or_manager_required
def add_equipment_page():
    return render_template('add_equipment.html')

@app.route('/chat')
@login_required
def chat():
    return render_template('chat.html')

# ==================== API - HOTEL SETTINGS ====================

@app.route('/api/hotel-info')
@login_required
def api_get_hotel_info():
    return jsonify(get_hotel_info())

@app.route('/api/hotel-settings', methods=['PUT'])
@login_required
@admin_required
def update_hotel_settings():
    data = request.get_json()
    try:
        settings = HotelSettings.query.first()
        if not settings:
            settings = HotelSettings()
            db.session.add(settings)
        settings.hotel_name = data.get('hotel_name', settings.hotel_name)
        settings.city = data.get('city', settings.city)
        settings.province = data.get('province', settings.province)
        settings.address = data.get('address', settings.address)
        settings.phone = data.get('phone', settings.phone)
        settings.email = data.get('email', settings.email)
        settings.website = data.get('website', settings.website)
        settings.star_rating = int(data['star_rating']) if data.get('star_rating') else settings.star_rating
        settings.total_rooms = int(data['total_rooms']) if data.get('total_rooms') else settings.total_rooms
        settings.hotel_type = data.get('hotel_type', settings.hotel_type)
        settings.updated_by = session.get('user_id')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Hotel settings updated successfully', 'hotel': get_hotel_info()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - DASHBOARD ====================

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    hotel = get_hotel_info()
    total_equipment = Equipment.query.filter_by(status='ACTIVE').count()
    high_risk = medium_risk = low_risk = 0
    for eq in Equipment.query.filter_by(status='ACTIVE').all():
        risk_level, _ = predict_risk_simple(eq)
        if risk_level == 'High': high_risk += 1
        elif risk_level == 'Medium': medium_risk += 1
        else: low_risk += 1
    today = datetime.now().date()
    overdue_count = Equipment.query.filter(Equipment.next_service_date < today, Equipment.status == 'ACTIVE').count()
    unread_alerts = Alert.query.filter_by(is_read=False).count()
    return jsonify({'total_equipment': total_equipment, 'high_risk': high_risk, 'medium_risk': medium_risk, 'low_risk': low_risk, 'overdue_count': overdue_count, 'unread_alerts': unread_alerts, 'hotel_name': hotel['hotel_name']})

@app.route('/api/risk-analysis')
@login_required
def risk_analysis():
    days = request.args.get('days', 7, type=int)
    if days > 90: days = 90
    if days < 1: days = 7
    start_date = datetime.now() - timedelta(days=days)
    predictions = Prediction.query.filter(Prediction.prediction_date >= start_date).all()
    data = {}
    for pred in predictions:
        date_str = pred.prediction_date.strftime('%Y-%m-%d')
        if date_str not in data:
            data[date_str] = {'Low': 0, 'Medium': 0, 'High': 0}
        risk_key = pred.risk_level.replace(' Risk', '').replace('Risk', '').strip()
        if risk_key in data[date_str]:
            data[date_str][risk_key] += 1
    return jsonify(data)

@app.route('/api/recent-predictions')
@login_required
def recent_predictions():
    predictions = Prediction.query.order_by(Prediction.prediction_date.desc()).limit(10).all()
    result = []
    for pred in predictions:
        result.append({'id': pred.id, 'user': pred.user.username if pred.user else 'System', 'date': pred.prediction_date.strftime('%Y-%m-%d'), 'risk_level': pred.risk_level, 'risk_score': pred.risk_score, 'summary': pred.summary, 'equipment_type': pred.equipment.equipment_type if pred.equipment else 'N/A', 'location': pred.equipment.location if pred.equipment else 'N/A'})
    return jsonify(result)

# ==================== API - EQUIPMENT ====================

@app.route('/api/equipment')
@login_required
def get_equipment():
    equipment_list = Equipment.query.filter_by(status='ACTIVE').all()
    result = []
    for eq in equipment_list:
        risk_level, risk_score = predict_risk_simple(eq)
        result.append({'id': eq.id, 'equipment_id': eq.equipment_id, 'equipment_type': eq.equipment_type, 'serial_number': eq.serial_number, 'manufacturer': eq.manufacturer, 'location': eq.location, 'floor': eq.floor, 'zone': eq.zone, 'last_service_date': eq.last_service_date.strftime('%Y-%m-%d') if eq.last_service_date else 'N/A', 'next_service_date': eq.next_service_date.strftime('%Y-%m-%d') if eq.next_service_date else 'N/A', 'risk_level': risk_level, 'risk_score': risk_score, 'condition': eq.condition, 'status': eq.status})
    return jsonify(result)

@app.route('/api/equipment/<int:equipment_id>')
@login_required
def get_equipment_details(equipment_id):
    eq = Equipment.query.get_or_404(equipment_id)
    risk_level, risk_score = predict_risk_simple(eq)
    age = datetime.now().year - eq.install_year
    maintenance_count = MaintenanceRecord.query.filter_by(equipment_id=equipment_id).count()
    days_since_service = (datetime.now().date() - eq.last_service_date).days if eq.last_service_date else 0
    return jsonify({'id': eq.id, 'equipment_id': eq.equipment_id, 'equipment_type': eq.equipment_type, 'serial_number': eq.serial_number, 'manufacturer': eq.manufacturer, 'model': eq.model, 'install_year': eq.install_year, 'age': age, 'last_service_date': eq.last_service_date.strftime('%Y-%m-%d') if eq.last_service_date else 'N/A', 'next_service_date': eq.next_service_date.strftime('%Y-%m-%d') if eq.next_service_date else 'N/A', 'service_interval_months': eq.service_interval_months, 'days_since_service': days_since_service, 'location': eq.location, 'floor': eq.floor, 'zone': eq.zone, 'near_emergency_exit': eq.near_emergency_exit, 'condition': eq.condition, 'status': eq.status, 'usage_level': eq.usage_level, 'humidity_level': eq.humidity_level, 'certification_status': eq.certification_status, 'last_audit_score': eq.last_audit_score, 'risk_level': risk_level, 'risk_score': risk_score, 'maintenance_count': maintenance_count, 'notes': eq.notes})

@app.route('/api/equipment/<int:equipment_id>/history')
@login_required
def get_equipment_history(equipment_id):
    records = MaintenanceRecord.query.filter_by(equipment_id=equipment_id).order_by(MaintenanceRecord.maintenance_date.desc()).all()
    result = []
    for record in records:
        result.append({'id': record.id, 'record_id': record.record_id, 'maintenance_date': record.maintenance_date.strftime('%Y-%m-%d'), 'maintenance_type': record.maintenance_type, 'technician_name': record.technician_name, 'technician_company': record.technician_company, 'issues_found': record.issues_found, 'actions_taken': record.actions_taken, 'parts_replaced': record.parts_replaced, 'cost_lkr': record.cost_lkr, 'duration_hours': record.duration_hours, 'status': record.status, 'notes': record.notes})
    return jsonify(result)

@app.route('/api/equipment', methods=['POST'])
@login_required
@admin_or_manager_required
def create_equipment():
    data = request.get_json()
    try:
        last_eq = Equipment.query.order_by(Equipment.id.desc()).first()
        new_id = f"FE{str(int(last_eq.equipment_id[2:]) + 1).zfill(6)}" if last_eq and last_eq.equipment_id else "FE000001"
        install_date = datetime.strptime(data['install_date'], '%Y-%m-%d').date() if data.get('install_date') else None
        last_service = datetime.strptime(data['last_service_date'], '%Y-%m-%d').date()
        next_service = last_service + timedelta(days=int(data.get('service_interval_months', 6)) * 30)
        new_equipment = Equipment(equipment_id=new_id, equipment_type=data['equipment_type'], serial_number=data.get('serial_number'), manufacturer=data.get('manufacturer'), model=data.get('model'), install_year=int(data['install_year']), install_date=install_date, last_service_date=last_service, next_service_date=next_service, service_interval_months=int(data.get('service_interval_months', 6)), location=data.get('location'), floor=data.get('floor'), zone=data.get('zone'), near_emergency_exit=data.get('near_emergency_exit', False), condition=data.get('condition', 'Good'), usage_level=data.get('usage_level', 'Medium'), daily_exposure_hours=int(data['daily_exposure_hours']) if data.get('daily_exposure_hours') else None, humidity_level=data.get('humidity_level'), temperature_avg=int(data['temperature_avg']) if data.get('temperature_avg') else None, coastal_exposure=data.get('coastal_exposure', False), certification_status=data.get('certification_status', 'Valid'), last_audit_score=int(data['last_audit_score']) if data.get('last_audit_score') else None, notes=data.get('notes'), created_by=session.get('user_id'))
        db.session.add(new_equipment)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Equipment registered successfully', 'equipment_id': new_equipment.id, 'equipment_code': new_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - MAINTENANCE ====================

@app.route('/api/maintenance', methods=['POST'])
@login_required
def add_maintenance_record():
    data = request.get_json()
    try:
        last_record = MaintenanceRecord.query.order_by(MaintenanceRecord.id.desc()).first()
        new_id = f"M{str(int(last_record.record_id[1:]) + 1).zfill(6)}" if last_record and last_record.record_id else "M000001"
        maintenance_date = datetime.strptime(data['maintenance_date'], '%Y-%m-%d').date()
        next_service = datetime.strptime(data['next_service_date'], '%Y-%m-%d').date() if data.get('next_service_date') else None
        record = MaintenanceRecord(record_id=new_id, equipment_id=int(data['equipment_id']), user_id=session.get('user_id'), maintenance_date=maintenance_date, maintenance_type=data.get('maintenance_type'), technician_name=data.get('technician_name'), technician_company=data.get('technician_company'), issues_found=data.get('issues_found'), actions_taken=data.get('actions_taken'), parts_replaced=data.get('parts_replaced'), cost_lkr=float(data['cost_lkr']) if data.get('cost_lkr') else None, duration_hours=float(data['duration_hours']) if data.get('duration_hours') else None, next_service_date=next_service, status=data.get('status', 'Completed'), notes=data.get('notes'))
        db.session.add(record)
        equipment = Equipment.query.get(int(data['equipment_id']))
        if equipment:
            equipment.last_service_date = maintenance_date
            if next_service:
                equipment.next_service_date = next_service
        db.session.commit()
        return jsonify({'success': True, 'message': 'Maintenance record added', 'record_id': new_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/maintenance/<int:record_id>', methods=['PUT'])
@login_required
@admin_required
def update_maintenance_record(record_id):
    data = request.get_json()
    try:
        record = MaintenanceRecord.query.get_or_404(record_id)
        if data.get('maintenance_date'):
            record.maintenance_date = datetime.strptime(data['maintenance_date'], '%Y-%m-%d').date()
        record.maintenance_type = data.get('maintenance_type', record.maintenance_type)
        record.technician_name = data.get('technician_name', record.technician_name)
        record.technician_company = data.get('technician_company', record.technician_company)
        record.issues_found = data.get('issues_found', record.issues_found)
        record.actions_taken = data.get('actions_taken', record.actions_taken)
        record.parts_replaced = data.get('parts_replaced', record.parts_replaced)
        if data.get('cost_lkr'):
            record.cost_lkr = float(data['cost_lkr'])
        if data.get('duration_hours'):
            record.duration_hours = float(data['duration_hours'])
        record.notes = data.get('notes', record.notes)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Maintenance record updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/maintenance/<int:record_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_maintenance_record(record_id):
    try:
        record = MaintenanceRecord.query.get_or_404(record_id)
        record_id_for_log = record.record_id
        db.session.delete(record)
        db.session.commit()
        print(f"✅ Maintenance record {record_id_for_log} deleted by admin")
        return jsonify({'success': True, 'message': 'Maintenance record deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - PREDICTIONS ====================

@app.route('/api/predict', methods=['POST'])
@login_required
def predict():
    data = request.get_json()
    equipment_id = data.get('equipment_id')
    try:
        equipment = Equipment.query.get_or_404(equipment_id)
        risk_level, risk_score = predict_risk_simple(equipment)
        age = datetime.now().year - equipment.install_year
        summary = f"Risk Score: {risk_score}% | Age: {age} years | Condition: {equipment.condition}"
        prediction = Prediction(equipment_id=equipment.id, user_id=session.get('user_id'), risk_level=risk_level, risk_score=risk_score, summary=summary)
        db.session.add(prediction)
        db.session.commit()
        return jsonify({'success': True, 'risk_level': risk_level, 'risk_score': risk_score, 'summary': summary})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/predictions')
@login_required
def get_all_predictions():
    predictions = Prediction.query.order_by(Prediction.prediction_date.desc()).all()
    result = []
    for pred in predictions:
        result.append({
            'id': pred.id, 
            'user': pred.user.username if pred.user else 'System', 
            'date': pred.prediction_date.strftime('%Y-%m-%d'), 
            'risk_level': pred.risk_level, 
            'risk_score': pred.risk_score, 
            'summary': pred.summary, 
            'equipment_type': pred.equipment.equipment_type if pred.equipment else 'N/A', 
            'equipment_id': pred.equipment.equipment_id if pred.equipment else 'N/A',
            'location': pred.equipment.location if pred.equipment else 'N/A'
        })
    return jsonify(result)

# ==================== API - ALERTS ====================

@app.route('/api/alerts')
@login_required
def get_alerts():
    alerts = Alert.query.order_by(Alert.created_at.desc()).limit(50).all()
    result = []
    for alert in alerts:
        equipment = None
        equipment_id_code = None
        equipment_db_id = None
        read_by_user = None
        resolved_by_user = None
        
        if alert.equipment_id:
            equipment = db.session.get(Equipment, alert.equipment_id)
            if equipment:
                equipment_id_code = equipment.equipment_id
                equipment_db_id = equipment.id
        
        if alert.read_by:
            user = db.session.get(User, alert.read_by)
            read_by_user = user.username if user else None
            
        if alert.resolved_by:
            user = db.session.get(User, alert.resolved_by)
            resolved_by_user = user.username if user else None
        
        result.append({
            'id': alert.id,
            'alert_type': alert.alert_type,
            'severity': alert.severity,
            'message': alert.message,
            'created_at': alert.created_at.strftime('%Y-%m-%d %H:%M'),
            'is_read': alert.is_read,
            'is_resolved': alert.is_resolved,
            'read_at': alert.read_at.strftime('%Y-%m-%d %H:%M') if alert.read_at else None,
            'read_by': read_by_user,
            'resolved_at': alert.resolved_at.strftime('%Y-%m-%d %H:%M') if alert.resolved_at else None,
            'resolved_by': resolved_by_user,
            'equipment_type': alert.equipment.equipment_type if alert.equipment else 'N/A',
            'location': alert.equipment.location if alert.equipment else 'N/A',
            'equipment_id': equipment_id_code,
            'equipment_db_id': equipment_db_id
        })
    return jsonify(result)

@app.route('/api/alerts/<int:alert_id>/mark-read', methods=['POST'])
@login_required
def mark_alert_read(alert_id):
    alert = Alert.query.get_or_404(alert_id)
    alert.is_read = True
    alert.read_at = datetime.utcnow()
    alert.read_by = session.get('user_id')
    db.session.commit()
    return jsonify({'success': True, 'message': 'Alert marked as read'})

@app.route('/api/alerts/<int:alert_id>/resolve', methods=['POST'])
@login_required
@admin_or_manager_required
def resolve_alert(alert_id):
    alert = Alert.query.get_or_404(alert_id)
    alert.is_resolved = True
    alert.resolved_at = datetime.utcnow()
    alert.resolved_by = session.get('user_id')
    db.session.commit()
    return jsonify({'success': True, 'message': 'Alert resolved successfully'})

@app.route('/api/check-alerts')
@login_required
def check_alerts():
    try:
        alerts_created = 0
        equipment_list = Equipment.query.filter_by(status='ACTIVE').all()
        
        for eq in equipment_list:
            try:
                if not eq.install_year:
                    continue
                    
                risk_level, risk_score = predict_risk_simple(eq)
                
                if risk_level == 'High':
                    existing_alert = Alert.query.filter_by(
                        equipment_id=eq.id, 
                        is_resolved=False, 
                        alert_type='High Risk Equipment'
                    ).first()
                    
                    if not existing_alert:
                        alert = Alert(
                            equipment_id=eq.id, 
                            alert_type='High Risk Equipment', 
                            severity='High', 
                            message=f'{eq.equipment_type} at {eq.location} has HIGH RISK (Score: {risk_score}%). Immediate attention required!'
                        )
                        db.session.add(alert)
                        alerts_created += 1
                
                if eq.next_service_date:
                    days_overdue = (datetime.now().date() - eq.next_service_date).days
                    if days_overdue > 0:
                        existing_overdue = Alert.query.filter_by(
                            equipment_id=eq.id, 
                            is_resolved=False, 
                            alert_type='Service Overdue'
                        ).first()
                        
                        if not existing_overdue:
                            if days_overdue > 30:
                                severity = 'Critical'
                            elif days_overdue > 7:
                                severity = 'High'
                            else:
                                severity = 'Medium'
                                
                            alert = Alert(
                                equipment_id=eq.id, 
                                alert_type='Service Overdue', 
                                severity=severity, 
                                message=f'{eq.equipment_type} at {eq.location} is {days_overdue} days overdue for service!'
                            )
                            db.session.add(alert)
                            alerts_created += 1
                            
            except Exception as eq_error:
                print(f"Error checking equipment {eq.id}: {str(eq_error)}")
                continue
        
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f'{alerts_created} new alerts created', 
            'new_alerts': alerts_created
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Alert check error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - EXPORT ====================

@app.route('/api/export/equipment')
@login_required
def export_equipment():
    hotel = get_hotel_info()
    equipment_list = Equipment.query.filter_by(status='ACTIVE').all()
    output = BytesIO()
    output.write(f'Equipment Report - {hotel["hotel_name"]}\n'.encode())
    output.write(b'Equipment ID,Type,Location,Floor,Last Service,Next Service,Risk Level,Condition\n')
    for eq in equipment_list:
        risk_level, _ = predict_risk_simple(eq)
        line = f"{eq.equipment_id},{eq.equipment_type},{eq.location},{eq.floor},"
        line += f"{eq.last_service_date.strftime('%Y-%m-%d') if eq.last_service_date else 'N/A'},"
        line += f"{eq.next_service_date.strftime('%Y-%m-%d') if eq.next_service_date else 'N/A'},"
        line += f"{risk_level},{eq.condition}\n"
        output.write(line.encode())
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True, download_name=f'{hotel["hotel_name"]}_equipment_{datetime.now().strftime("%Y%m%d")}.csv')

@app.route('/api/export/maintenance')
@login_required
def export_maintenance():
    hotel = get_hotel_info()
    records = MaintenanceRecord.query.all()
    output = BytesIO()
    output.write(f'Maintenance Report - {hotel["hotel_name"]}\n'.encode())
    output.write(b'Record ID,Equipment ID,Date,Type,Technician,Company,Cost LKR,Status\n')
    for rec in records:
        equipment = Equipment.query.get(rec.equipment_id)
        line = f"{rec.record_id},{equipment.equipment_id if equipment else 'N/A'},{rec.maintenance_date.strftime('%Y-%m-%d')},"
        line += f"{rec.maintenance_type},{rec.technician_name},{rec.technician_company},{rec.cost_lkr or 0},{rec.status}\n"
        output.write(line.encode())
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True, download_name=f'{hotel["hotel_name"]}_maintenance_{datetime.now().strftime("%Y%m%d")}.csv')

@app.route('/api/export/excel')
@login_required
def export_excel():
    """Export comprehensive Excel workbook with multiple sheets"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel export not available. Install openpyxl: pip install openpyxl'}), 500
    
    try:
        hotel = get_hotel_info()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # ===== SHEET 1: Summary =====
        ws_summary = wb.create_sheet("Summary")
        ws_summary.merge_cells('A1:D1')
        title_cell = ws_summary['A1']
        title_cell.value = f'{hotel["hotel_name"]} - Fire Safety Report'
        title_cell.font = Font(size=16, bold=True, color="FF6B35")
        title_cell.alignment = Alignment(horizontal='center')
        
        ws_summary['A3'] = 'Report Date:'
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws_summary['A4'] = 'Generated By:'
        ws_summary['B4'] = session.get('username', 'System')
        
        # Statistics
        equipment_list = Equipment.query.filter_by(status='ACTIVE').all()
        total = len(equipment_list)
        high_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'High')
        medium_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'Medium')
        low_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'Low')
        overdue = sum(1 for eq in equipment_list if eq.next_service_date and eq.next_service_date < datetime.now().date())
        
        ws_summary['A6'] = 'STATISTICS'
        ws_summary['A6'].font = Font(bold=True, size=12)
        stats_data = [
            ['Total Equipment', total],
            ['High Risk Items', high_risk],
            ['Medium Risk Items', medium_risk],
            ['Low Risk Items', low_risk],
            ['Overdue Service', overdue]
        ]
        for idx, (label, value) in enumerate(stats_data, start=7):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'B{idx}'].font = Font(bold=True)
        
        # ===== SHEET 2: Equipment Inventory =====
        ws_equip = wb.create_sheet("Equipment Inventory")
        headers = ['Equipment ID', 'Type', 'Location', 'Floor', 'Install Year', 'Last Service', 'Next Service', 'Risk Level', 'Condition', 'Status']
        ws_equip.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        for cell in ws_equip[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for eq in equipment_list:
            risk_level, risk_score = predict_risk_simple(eq)
            ws_equip.append([
                eq.equipment_id,
                eq.equipment_type,
                eq.location,
                eq.floor,
                eq.install_year,
                eq.last_service_date.strftime('%Y-%m-%d') if eq.last_service_date else 'N/A',
                eq.next_service_date.strftime('%Y-%m-%d') if eq.next_service_date else 'N/A',
                risk_level,
                eq.condition,
                eq.status
            ])
        
        # Auto-size columns
        for column in ws_equip.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_equip.column_dimensions[column_letter].width = adjusted_width
        
        # ===== SHEET 3: Maintenance Log =====
        ws_maint = wb.create_sheet("Maintenance Log")
        headers = ['Record ID', 'Equipment ID', 'Date', 'Type', 'Technician', 'Company', 'Cost (LKR)', 'Duration (hrs)', 'Status']
        ws_maint.append(headers)
        
        # Style header
        for cell in ws_maint[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add maintenance records
        records = MaintenanceRecord.query.order_by(MaintenanceRecord.maintenance_date.desc()).all()
        for rec in records:
            equipment = Equipment.query.get(rec.equipment_id)
            ws_maint.append([
                rec.record_id,
                equipment.equipment_id if equipment else 'N/A',
                rec.maintenance_date.strftime('%Y-%m-%d'),
                rec.maintenance_type,
                rec.technician_name,
                rec.technician_company,
                rec.cost_lkr or 0,
                rec.duration_hours or 0,
                rec.status
            ])
        
        # Auto-size columns
        for column in ws_maint.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_maint.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'{hotel["hotel_name"]}_FireSafety_Report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/pdf')
@login_required
def export_pdf():
    """Export professional PDF compliance report"""
    if not PDF_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF export not available. Install reportlab: pip install reportlab'}), 500
    
    try:
        hotel = get_hotel_info()
        output = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FF6B35'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#FF6B35'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph(f'{hotel["hotel_name"]}', title_style))
        story.append(Paragraph('Fire Safety Compliance Report', styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        # Report Info
        report_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        story.append(Paragraph(f'<b>Report Date:</b> {report_date}', styles['Normal']))
        story.append(Paragraph(f'<b>Generated By:</b> {session.get("username", "System")}', styles['Normal']))
        story.append(Paragraph(f'<b>Location:</b> {hotel["city"]}, {hotel["province"]}', styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph('Executive Summary', heading_style))
        
        equipment_list = Equipment.query.filter_by(status='ACTIVE').all()
        total = len(equipment_list)
        high_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'High')
        medium_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'Medium')
        low_risk = sum(1 for eq in equipment_list if predict_risk_simple(eq)[0] == 'Low')
        overdue = sum(1 for eq in equipment_list if eq.next_service_date and eq.next_service_date < datetime.now().date())
        
        summary_data = [
            ['Metric', 'Count'],
            ['Total Equipment', str(total)],
            ['High Risk Items', str(high_risk)],
            ['Medium Risk Items', str(medium_risk)],
            ['Low Risk Items', str(low_risk)],
            ['Overdue for Service', str(overdue)]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Equipment Inventory
        story.append(Paragraph('Equipment Inventory', heading_style))
        
        equip_data = [['ID', 'Type', 'Location', 'Risk', 'Condition']]
        for eq in equipment_list[:20]:  # Limit to 20 for PDF
            risk_level, _ = predict_risk_simple(eq)
            equip_data.append([
                eq.equipment_id,
                eq.equipment_type[:20],
                eq.location[:15],
                risk_level,
                eq.condition
            ])
        
        equip_table = Table(equip_data, colWidths=[1*inch, 2*inch, 1.5*inch, 1*inch, 1*inch])
        equip_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(equip_table)
        
        if len(equipment_list) > 20:
            story.append(Paragraph(f'<i>Showing 20 of {total} items. Full list available in Excel export.</i>', styles['Normal']))
        
        story.append(PageBreak())
        
        # Maintenance Summary
        story.append(Paragraph('Recent Maintenance Activity', heading_style))
        
        records = MaintenanceRecord.query.order_by(MaintenanceRecord.maintenance_date.desc()).limit(15).all()
        maint_data = [['Date', 'Equipment', 'Type', 'Technician', 'Status']]
        for rec in records:
            equipment = Equipment.query.get(rec.equipment_id)
            maint_data.append([
                rec.maintenance_date.strftime('%Y-%m-%d'),
                equipment.equipment_id if equipment else 'N/A',
                rec.maintenance_type[:20],
                rec.technician_name[:15],
                rec.status
            ])
        
        maint_table = Table(maint_data, colWidths=[1*inch, 1.2*inch, 1.8*inch, 1.5*inch, 1*inch])
        maint_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(maint_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Compliance Notes
        story.append(Paragraph('Compliance Notes', heading_style))
        compliance_text = f'''
        This report provides a comprehensive overview of fire safety equipment at {hotel["hotel_name"]}. 
        All equipment should be maintained according to manufacturer specifications and local fire safety regulations.
        High-risk items require immediate attention. Medium-risk items should be scheduled for service within 30 days.
        Regular inspections and maintenance are critical for ensuring occupant safety and regulatory compliance.
        '''
        story.append(Paragraph(compliance_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        filename = f'{hotel["hotel_name"]}_Compliance_Report_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"PDF export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - USERS ====================

@app.route('/api/users')
@login_required
@admin_or_manager_required
def get_users():
    users = User.query.all()
    return jsonify([{'id': u.id, 'username': u.username, 'role': u.role, 'email': u.email, 'phone': u.phone, 'date_joined': u.date_joined.strftime('%Y-%m-%d'), 'status': u.status} for u in users])

@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def create_user():
    data = request.get_json()
    try:
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'success': False, 'error': 'Username already exists'}), 400
        
        # Hash the password with bcrypt
        hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        new_user = User(
            username=data['username'],
            password=hashed_password,  # Store hashed password
            role=data['role'],
            email=data['email'],
            phone=data.get('phone')
        )
        db.session.add(new_user)
        db.session.commit()
        return jsonify({'success': True, 'message': 'User created successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def update_user(user_id):
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        user.username = data.get('username', user.username)
        user.role = data.get('role', user.role)
        user.email = data.get('email', user.email)
        user.phone = data.get('phone', user.phone)
        user.status = data.get('status', user.status)
        if data.get('password'):
            # Hash the password with bcrypt
            user.password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        db.session.commit()
        return jsonify({'success': True, 'message': 'User updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(user_id):
    try:
        if user_id == session.get('user_id'):
            return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400
        
        user = User.query.get_or_404(user_id)
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API - CHAT SYSTEM ====================

@app.route('/api/chat/rooms/create', methods=['POST'])
@login_required
@admin_or_manager_required
def create_chat_room():
    """Create new chat room - Admin and Manager only"""
    data = request.get_json()
    
    room = ChatRoom(
        name=data.get('name'),
        room_type=data.get('room_type', 'general'),
        equipment_id=data.get('equipment_id'),
        created_by=session.get('user_id')
    )
    
    db.session.add(room)
    db.session.commit()
    
    return jsonify({'success': True, 'room_id': room.id, 'message': 'Chat room created successfully'})


@app.route('/api/chat/rooms/<int:room_id>/delete', methods=['DELETE'])
@login_required
@admin_required
def delete_chat_room(room_id):
    """Delete chat room - Admin only"""
    try:
        room = ChatRoom.query.get_or_404(room_id)
        room_name = room.name
        
        # Mark as inactive instead of deleting (preserves chat history)
        room.is_active = False
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Chat room "{room_name}" deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/chat/rooms')
@login_required
def get_chat_rooms():
    """Get all active chat rooms with user permissions"""
    rooms = ChatRoom.query.filter_by(is_active=True).order_by(ChatRoom.created_at.desc()).all()
    
    rooms_data = []
    current_user_role = session.get('role')
    
    for room in rooms:
        last_message = ChatMessage.query.filter_by(room_id=room.id).order_by(ChatMessage.created_at.desc()).first()
        
        if last_message:
            unread_count = ChatMessage.query.filter(
                ChatMessage.room_id == room.id,
                ChatMessage.user_id != session.get('user_id'),
                ~ChatMessage.read_by.any(ChatReadReceipt.user_id == session.get('user_id'))
            ).count()
        else:
            unread_count = 0
        
        # Get creator info
        creator = User.query.get(room.created_by) if room.created_by else None
        
        rooms_data.append({
            'id': room.id,
            'name': room.name,
            'room_type': room.room_type,
            'equipment_id': room.equipment_id,
            'created_by': creator.username if creator else 'System',
            'can_delete': current_user_role == 'Admin',
            'last_message': {
                'text': last_message.message[:50] + '...' if last_message and len(last_message.message) > 50 else last_message.message if last_message else None,
                'user': last_message.user.username if last_message else None,
                'time': last_message.created_at.isoformat() if last_message else None
            } if last_message else None,
            'unread_count': unread_count
        })
    
    return jsonify({
        'success': True, 
        'rooms': rooms_data,
        'user_role': current_user_role,
        'can_create': current_user_role in ['Admin', 'Manager']
    })


@app.route('/api/chat/rooms/<int:room_id>/messages')
@login_required
def get_room_messages(room_id):
    """Get all messages for a chat room"""
    room = ChatRoom.query.get_or_404(room_id)
    messages = ChatMessage.query.filter_by(room_id=room_id).order_by(ChatMessage.created_at.asc()).all()
    
    messages_data = []
    for msg in messages:
        read_receipt = ChatReadReceipt.query.filter_by(
            message_id=msg.id,
            user_id=session.get('user_id')
        ).first()
        
        # Get all users who have read this message
        read_by_list = []
        for receipt in msg.read_by:
            if receipt.user_id != msg.user_id:  # Exclude message sender
                user = User.query.get(receipt.user_id)
                if user:
                    read_by_list.append({
                        'username': user.username,
                        'read_at': receipt.read_at.isoformat()
                    })
        
        messages_data.append({
            'id': msg.id,
            'user_id': msg.user_id,
            'username': msg.user.username,
            'message': msg.message,
            'attachment_url': msg.attachment_url,
            'attachment_name': msg.attachment_name,
            'created_at': msg.created_at.isoformat(),
            'is_edited': msg.is_edited,
            'edited_at': msg.edited_at.isoformat() if msg.edited_at else None,
            'is_mine': msg.user_id == session.get('user_id'),
            'is_read': read_receipt is not None,
            'read_by': read_by_list  # New field
        })
    
    return jsonify({
        'success': True,
        'room': {
            'id': room.id,
            'name': room.name,
            'room_type': room.room_type
        },
        'messages': messages_data
    })


@app.route('/api/chat/messages/send', methods=['POST'])
@login_required
def send_message():
    """Send a new message to a chat room"""
    data = request.get_json()
    
    message = ChatMessage(
        room_id=data.get('room_id'),
        user_id=session.get('user_id'),
        message=data.get('message'),
        attachment_url=data.get('attachment_url'),
        attachment_name=data.get('attachment_name')
    )
    
    db.session.add(message)
    db.session.commit()
    
    return jsonify({'success': True, 'message_id': message.id})


@app.route('/api/chat/messages/<int:message_id>/read', methods=['POST'])
@login_required
def mark_message_read(message_id):
    existing = ChatReadReceipt.query.filter_by(
        message_id=message_id,
        user_id=session.get('user_id')
    ).first()
    
    if not existing:
        receipt = ChatReadReceipt(
            message_id=message_id,
            user_id=session.get('user_id')
        )
        db.session.add(receipt)
        db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/chat/unread-count')
@login_required
def get_unread_count():
    unread = ChatMessage.query.filter(
        ChatMessage.user_id != session.get('user_id'),
        ~ChatMessage.read_by.any(ChatReadReceipt.user_id == session.get('user_id'))
    ).count()
    
    return jsonify({'success': True, 'unread_count': unread})

# ==================== INITIALIZE ====================

def init_db():
    with app.app_context():
        db.create_all()
        
        if HotelSettings.query.count() == 0:
            default_settings = HotelSettings(hotel_name='Amaya Hills', city='kandy', province='Central', star_rating=5, total_rooms=100, hotel_type='City Hotel')
            db.session.add(default_settings)
            db.session.commit()
            print("✅ Default hotel settings created!")
        
        # Create default users FIRST (before chat rooms need them)
        create_default_users()
        
        # Create default chat rooms
        create_default_chat_rooms()
        
        print("✅ Database initialized successfully!")

if __name__ == '__main__':
    init_db()
    with app.app_context():
        hotel = get_hotel_info()
        print("=" * 70)
        print(f"🔥 FireSense - {hotel['hotel_name']}")
        print("=" * 70)
        print("✅ Using PostgreSQL Database")
        print("✅ Role-Based Access Control Enabled")
        print("=" * 70)
        print("👑 ADMIN: Full access to all features")
        print("👔 MANAGER: View users, resolve alerts, add equipment/maintenance")
        print("🔧 TECHNICIAN: View only, mark alerts as read")
        print("=" * 70)
        print(f"📍 Hotel: {hotel['hotel_name']}")
        print(f"📍 Location: {hotel.get('city', 'N/A')}, {hotel.get('province', 'N/A')}")
        print("=" * 70)
        print("🌐 Dashboard: http://localhost:5000")
        print("👤 Login: admin / admin123")
        print("💬 Chat: http://localhost:5000/chat")
        print("=" * 70)
    app.run(debug=True, host='0.0.0.0', port=5000)
